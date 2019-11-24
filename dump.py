import xlsxwriter
from collections import OrderedDict
from openpyxl import load_workbook

SPECIAL_CHARACTERS = {
    'ō': '[o]',
    'Ō': '[O]',
    'ū': '[u]',
    '\uff0d': '\u30fc',   # Katakana long dash mark fix
    #'\u2161': '\x87\x55', # II Ⅱ fix
}

def unpack(s, t=None):
    if t is None:
        t = str(s)[2:]
        s = str(s)[0:2]
    s = int(s, 16)
    t = int(t, 16)
    value = (t * 0x100) + s
    return value


def pack(h):
    s = h % 0x100
    t = h // 0x100
    return (s, t)


def ascii_to_hex_string(eng, control_codes={}):
    """Returns a hex string of the ascii bytes of a given english string."""
    eng_bytestring = ""
    if not eng:
        return ""
    else:
        try:
            eng = str(eng)
        except UnicodeEncodeError:
            # Tried to encode a fullwidth number. Encode it as sjis instead.
            eng = eng.encode('shift-jis')

        eng_bytestring = eng

        for cc in control_codes:
            cc_hex = ascii_to_hex_string(cc)
            if cc_hex in eng_bytestring:
                eng_bytestring = eng_bytestring.replace(cc_hex, control_codes[cc])

        return eng_bytestring


def sjis_to_hex_string(jp, control_codes={}):
    """Returns a hex string of the Shift JIS bytes of a given japanese string."""
    jp_bytestring = ""
    try:
        sjis = jp.encode('shift-jis')
    except AttributeError:
        # Trying to encode numbers throws an attribute error; they aren't important, so just keep the number
        sjis = str(jp)

    jp_bytestring = sjis

    for cc in control_codes:
        cc_hex = sjis_to_hex_string(cc)
        if cc_hex in jp_bytestring:
            jp_bytestring = jp_bytestring.replace(cc_hex, control_codes[cc])

    return jp_bytestring


class Translation(object):
    """Has an offset, a SJIS japanese string, and an ASCII english string."""
    def __init__(self, gamefile, location, japanese, english, category=None, portrait=None,
                 prefix=None, suffix=None, command=None, cd_location=None, compressed_location=None, total_location=None, pointers=None,
                 control_codes={}):
        self.location = location
        self.cd_location = cd_location
        self.compressed_location = compressed_location
        self.total_location = total_location
        self.gamefile = gamefile

        if prefix is not None:
            self.japanese = prefix + japanese
            self.english = prefix + english
        else:
            self.japanese = japanese
            self.english = english

        self.jp_bytestring = self.japanese
        self.en_bytestring = self.english

        self.category = category
        self.portrait = portrait
        self.prefix = prefix
        self.suffix = suffix
        self.pointers = pointers
        self.command = command

        for cc in control_codes:
            if cc in self.jp_bytestring:
                self.jp_bytestring = self.jp_bytestring.replace(cc, control_codes[cc])
            if cc in self.en_bytestring:
                self.en_bytestring = self.en_bytestring.replace(cc, control_codes[cc])


    def refresh_jp_bytestring(self):
        self.jp_bytestring = sjis_to_hex_string(self.japanese)
        return self.jp_bytestring

    def refresh_en_bytestring(self):
        self.en_bytestring = ascii_to_hex_string(self.english)
        return self.en_bytestring

    def __repr__(self):
        return "%s %s" % (hex(self.location), self.english)


class SegmentPointer:
    """Trying something different for LA."""
    def __init__(self, filestring, pointer_location, text_location):
        #self.segment = segment
        self.location = pointer_location
        self.filestring = filestring
        #self.location_in_segment = pointer_location - segment.start
        self.text_location = text_location
        #print(hex(self.location_in_segment))

    def edit(self, diff):
        #print("Editing %s with diff %s" % (self, diff))
        if diff == 0:
            return None
        #first = hex(self.segment.string[self.location_in_segment])
        #second = hex(self.segment.string[self.location_in_segment+1])
        first = hex(self.filestring[self.location])
        second = hex(self.filestring[self.location+1])

        old_value = unpack(first, second)
        new_value = old_value + diff

        new_bytes = new_value.to_bytes(length=2, byteorder='little')

        # TODO: Might want to store the new filestring somewhere? SOmewhere that other pointers are aware of
        #prefix = self.filestring[:self.location]
        #suffix = self.filestring[self.location+2:]

        #self.segment.string = prefix + new_bytes + suffix
        self.new_text_location = new_value
        #assert len(self.segment.string) == len(self.gamefile.original_filestring), (hex(len(self.segment.string)), hex(len(self.gamefile.original_filestring)))

        return new_bytes


class BorlandPointer(object):
    """Two-byte, little-endian pointer with a constant added to retrieve location."""
    def __init__(self, gamefile, pointer_location, text_location, separator=b'\x00'):
        # A BorlandPointer has only one location. The container OrderDicts have lists of pointers,
        # each having their own location.
        self.gamefile = gamefile
        self.constant = self.gamefile.pointer_constant
        self.location = pointer_location
        self.original_location = pointer_location

        self.original_text_location = text_location
        self.text_location = text_location
        self.separator = separator

        value_bytes = pack(self.original_text_location - self.constant)

        self.value = "%s %s" % ('{0:02x}'.format(value_bytes[0]), '{0:02x}'.format(value_bytes[1]))

    def text(self, control_codes={}):
        gamefile_slice = self.gamefile.filestring[self.text_location:self.text_location+30]
        gamefile_slice = gamefile_slice.split(self.separator)[0]
        
        for cc in control_codes:
            if cc == b'[BLANK]':
                continue
            gamefile_slice = gamefile_slice.replace(control_codes[cc], cc)

        try:
            gamefile_slice = gamefile_slice.decode('shift_jis')
        except:
            helpful_bytes = ' '.join(['{0:02x}'.format(b) for b in gamefile_slice])
            gamefile_slice = helpful_bytes
        return gamefile_slice

    def original_text(self):
        gamefile_slice = self.gamefile.original_filestring[self.text_location:self.text_location+45]
        gamefile_slice = gamefile_slice.split('\x00')[0]
        try:
            gamefile_slice = gamefile_slice.decode('shift_jis')
        except:
            helpful_bytes = ' '.join(['{0:02x}'.format(b) for b in gamefile_slice])
            gamefile_slice = helpful_bytes
        return gamefile_slice

    def move_pointer_location(self, diff):
        self.location += diff
        return self.location


    def edit(self, diff, block=None):
        print("Editing %s in file %s with diff %s. Block is %s" % (self, self.gamefile, diff, block))
        if block:
            # Let's edit the blockstring instead of the filestring
            #print("The pointer is at %s, the block is (%s, %s)" % (hex(self.location), hex(block.start), hex(block.stop)))
            b_location = self.location - block.start
            #print(hex(b_location))


            first = hex(block.blockstring[b_location])
            second = hex(block.blockstring[b_location+1])

            #print(first, second, hex(self.text_location + self.constant))

            old_value = unpack(first, second)
            new_value = old_value + diff

            new_bytes = new_value.to_bytes(length=2, byteorder='little')

            prefix = block.blockstring[:self.location - block.start]
            suffix = block.blockstring[self.location - block.start+2:]

            block.blockstring = prefix + new_bytes + suffix
            self.text_location = new_value + block.start

            return new_bytes

        else:
            first = hex(self.gamefile.filestring[self.location])
            second = hex(self.gamefile.filestring[self.location+1])
            #print(first, second)
            old_value = unpack(first, second)
            new_value = old_value + diff

            #print(hex(old_value))
            #print(diff)

            new_bytes = new_value.to_bytes(length=2, byteorder='little')
            print(hex(old_value), hex(new_value))
            print((first, second), repr(new_bytes))
            #new_first, new_second = bytearray(new_bytes[0]), bytearray(new_bytes[1])
            prefix = self.gamefile.filestring[:self.location]
            suffix = self.gamefile.filestring[self.location+2:]
            self.gamefile.filestring = prefix + new_bytes + suffix
            self.text_location = new_value


            #assert len(self.gamefile.filestring) == len(self.gamefile.original_filestring), (hex(len(self.gamefile.filestring)), hex(len(self.gamefile.original_filestring)))
            return new_bytes

    def __repr__(self):
        return "%s pointing to %s" % (hex(self.location), hex(self.text_location))

class PossessionerPointer(BorlandPointer):
    pass

class DumpExcel(object):
    """
    Takes a dump excel path, and lets you get a block's translations from it.
    """
    # TODO: Currently uses the order of the excel sheet. Might want to sort it by text_location in get_translations()...

    def __init__(self, path, control_codes={}):
        self.path = path
        self.workbook = load_workbook(self.path)
        self.control_codes = control_codes

        #self.rows_with_file = {}
        #print("About to cache rows with file")
        #self._cache_rows_with_file()


    def get_translations(self, target, sheet_name=None, include_blank=False, use_cd_location=False):
        """Get the translations for a file."""
        # Accepts a block, gamefile, or filename as "target".
        # If sheet_name is defined, the target will be the filenamne within that multi-file sheet.
        trans = []    # translations[offset] = Translation()

        if sheet_name:
            worksheet = self.workbook.get_sheet_by_name(sheet_name)
        else:
            try:
                worksheet = self.workbook.get_sheet_by_name(target.gamefile.filename)
            except KeyError:
                worksheet = self.workbook.get_sheet_by_name(target.gamefile.filename.lstrip('decompressed_').rstrip('.decompressed'))
            except AttributeError:
                try:
                    worksheet = self.workbook.get_sheet_by_name(target.filename)
                except AttributeError:
                    worksheet = self.workbook.get_sheet_by_name(target)

        first_row = list(worksheet.rows)[0]
        header_values = [t.value for t in first_row]

        try:
            offset_col = header_values.index('Offset (FD)')
            cd_offset_col = header_values.index('Offset (CD)')
            compressed_offset_col = None
        except ValueError:
            try:
                offset_col = header_values.index('Offset')
                cd_offset_col = None
                compressed_offset_col = header_values.index('Compressed Offset')
            except ValueError:
                offset_col = header_values.index('Offset')
                cd_offset_col = None
                compressed_offset_col = None

        jp_col = header_values.index('Japanese')

        # Appareden (and later games) have two (three?) English columns
        try:
            en_col = header_values.index('English (Typeset)')
        except ValueError:
            try:
                en_col = header_values.index('English (Ingame)')
            except ValueError:
                en_col = header_values.index('English')

        try:
            filename_col = header_values.index('File')
        except ValueError:
            try:
                filename_col = header_values.index('Filename')
            except ValueError:
                filename_col = None

        # TODO: These more ad-hoc columns might do better in a dictionary or something.

        try:
            category_col = header_values.index('Category')
        except ValueError:
            category_col = None

        try:
            portrait_col = header_values.index('Portrait')
        except ValueError:
            portrait_col = None

        try:
            suffix_col = header_values.index("Suffix")
        except ValueError:
            suffix_col = None

        try:
            prefix_col = header_values.index('Ctrl Codes')
        except ValueError:
            prefix_col = None

        try:
            command_col = header_values.index("Command")
        except ValueError:
            command_col = None

        try:
            pointer_col = header_values.index("Pointer")
        except ValueError:
            pointer_col = None

        try:
            total_offset_col = header_values.index("Offset (Total)")
        except ValueError:
            total_offset_col = None

        for row in list(worksheet.rows)[1:]:  # Skip the first row, it's just labels
            # Skip rows not for this block and file, if the target is a block
            if filename_col is not None:
                try:
                    start, stop = target.start, target.stop
                    if sheet_name and row[filename_col].value != target.gamefile.filename:
                        continue
                # Skip rows that aren't for this file, if a sheet name is specified
                except AttributeError:
                    if sheet_name and row[filename_col].value != target:
                        continue

            try:
                offset = int(row[offset_col].value, 16)
            except TypeError:
                # Either a blank line or a total value. Ignore it.
                offset = None

            try:
                cd_offset = int(row[cd_offset_col].value, 16)
            except TypeError:
                cd_offset = None

            try:
                compressed_offset = int(row[compressed_offset_col].value, 16)
            except TypeError:
                compressed_offset = None

            try:
                total_offset = int(row[total_offset_col].value, 16)
            except TypeError:
                total_offset = None

            if offset is None and cd_offset is None and compressed_offset is None:
                break

            try:
                start, stop = target.start, target.stop
                #print(start, stop)
                if use_cd_location:
                    offset = cd_offset
                if offset is None:
                    continue
                if not (target.start <= offset < target.stop):
                    continue
            except AttributeError:
                pass
            #print("Made it this far")

            if row[en_col].value is None and not include_blank:
                continue
            #print(sheet_name)
            #for sc in SPECIAL_CHARACTERS:
            #    japanese = japanese.replace(sc, SPECIAL_CHARACTERS[sc])
            #japanese = japanese.encode('shift-jis')
            if row[jp_col].value is None:
                japanese = b""
            else:
                japanese = row[jp_col].value.encode('shift-jis')
            #print(japanese)
            #print(row[en_col].value)
            if row[en_col].value is None:
                english = b""
            else:
                try:
                    english = row[en_col].value.encode('shift-jis')
                except AttributeError:   # Int column values
                    
                    english = str(row[en_col].value).encode('shift-jis')
                except UnicodeEncodeError:
                    print(hex(offset))
                    english = row[en_col].value
                    for ch in SPECIAL_CHARACTERS:
                        english = english.replace(ch, SPECIAL_CHARACTERS[ch])
                    english = english.encode('shift-jis')

            # Category, for Appareden equipment
            if category_col is not None:
                category = row[category_col].value
            else:
                category = None

            # Portrait ID, for Appareden dialogue
            if portrait_col is not None:
                portrait = row[portrait_col].value
            else:
                portrait = None

            # Suffix control codes, for Different Realm control code soup
            if suffix_col is not None:
                suffix = row[suffix_col].value
            else:
                suffix = None

            # Prefix control codes, for (secret project) control code soup
            if prefix_col is not None:
                if row[prefix_col].value:
                    prefix = row[prefix_col].value.encode('shift-jis')
                else:
                    prefix = None
            else:
                prefix = None

            # Command context, for (secret project) dialogue context
            if command_col is not None:
                command = row[command_col].value
            else:
                command = None

            # Pointer column, for Last Armageddon
            if pointer_col is not None:
                if row[pointer_col].value is not None:
                    pointers = [int(loc, 16) for loc in row[pointer_col].value.split("; ")]
                else:
                    pointers = None
            else:
                pointers=None


            # if isinstance(japanese, long):
            #    # Causes some encoding problems? Trying to skip them for now
            #    continue

            # Blank strings are None (non-iterable), so use "" instead.
            if not english:
                english = b""

            trans.append(Translation(target, offset, japanese, english,
                                     category=category, portrait=portrait,
                                     control_codes=self.control_codes,
                                     cd_location=cd_offset, compressed_location=compressed_offset, total_location=total_offset,
                                     suffix=suffix, prefix=prefix, 
                                     command=command, pointers=pointers
                                     ))
        return trans


class PointerExcel(object):
    def __init__(self, path):
        self.path = path
        try:
            self.workbook = load_workbook(self.path)
        except IOError:
            self.workbook = xlsxwriter.Workbook(self.path)

    def add_worksheet(self, title):
        self.worksheet = self.workbook.add_worksheet(title)
        sheet_format = {'bold': True,
                        'align': 'center',
                        'bottom': True,
                        'bg_color': 'gray'}
        header = self.workbook.add_format(sheet_format)
        self.worksheet.write(0, 0, 'Text Loc', header)
        self.worksheet.write(0, 1, 'Ptr Loc', header)
        self.worksheet.write(0, 2, 'Bytes', header)
        self.worksheet.write(0, 3, 'Points To', header)
        self.worksheet.write(0, 4, 'Comments', header)
        self.worksheet.set_column('A:A', 9)
        self.worksheet.set_column('B:B', 9)
        self.worksheet.set_column('C:C', 30)
        self.worksheet.set_column('D:D', 30)
        return self.worksheet

    def get_pointers(self, gamefile, pointer_sheet_name):
        pointers = OrderedDict()
        print(pointer_sheet_name)
        try:
            ws = self.workbook[pointer_sheet_name]
        except KeyError:
            # Sheet does not exist. Return an empty pointers list.
            print('No pointers for', pointer_sheet_name)
            return pointers
        except TypeError:
            print("Workbook object is not subscriptable...?")
            return pointers

        for i, row in enumerate(ws):
            if i == 0:
                continue
            text_location = int(row[0].value, 16)
            try:
                pointer_location = int(row[1].value, 16)
            except ValueError:
                print("Pointer with text location %s had no pointer location. Proceed with caution" % hex(text_location))
                continue
            ptr = BorlandPointer(gamefile, pointer_location, text_location)
            if text_location in pointers:
                pointers[text_location].append(ptr)
            else:
                pointers[text_location] = [ptr, ]
        return pointers

    def close(self):
        self.workbook.close()
"""
class DumpGoogleSheet(object):
"""
       # A dump in a Google Sheet, with methods to update various columns
"""
    def __init__(self, filename):
        gc = pygsheets.authorize(outh_file='client_secret_1010652634407-2d4gjkn44a5020jg6tl4hqjld1130fjs.apps.googleusercontent.com.json', no_cache=True)
        self.workbook = gc.open(filename)

def update_google_sheets(local_filename, google_filename):
    local = DumpExcel(local_filename)
    google = DumpGoogleSheet(google_filename)

    for name in local.workbook.get_sheet_names():
        local_worksheet = local.workbook.get_sheet_by_name(name)

        first_row = list(local_worksheet.rows)[0]
        header_values = [t.value for t in first_row]

        offset_col = header_values.index('Offset')
        jp_col = header_values.index('Japanese')
        # TODO: Probably want to update these too

        original_en_col = header_values.index('English (kuoushi)')
        ingame_en_col = header_values.index('English (Ingame)')

        google_worksheet = google.workbook.worksheet_by_title(name)

        google_values = google_worksheet.get_all_values(returnas='cell', include_empty=False)


        #print(google_values)

        # TODO: Propagate a deleted row from local to Google sheet; not supported yet

        # TODO: Sys dump has one less row in google sheet than local, but msg dump has same row count!
        print(len(google_values)-1, local_worksheet.max_row)
        assert len(google_values)-1 <= local_worksheet.max_row, "Row has been deleted, update manually"
        if len(google_values)-1 < local_worksheet.max_row:
            print("Row added to local sheet, inserting that in google sheet")
            for i, row in enumerate(local_worksheet):
                google_offset = google_values[i][0].value
                local_offset = row[0].value
                if google_offset != local_offset:
                    vals_list = [cell.value for cell in row]
                    for cell_i in range(len(vals_list)):
                        if vals_list[cell_i] is None:
                            vals_list[cell_i] = ''
                    google_worksheet.insert_rows(row=i, number=1, values=vals_list)


        # TODO: Want to update each direction separately...
        for en_col in (original_en_col, ingame_en_col):

            for i, row in enumerate(local_worksheet):
                try:
                    google_val = google_values[i][en_col].value
                except IndexError:
                    # Likely to be a sheet that doesn't have any values
                    print("Skipping this sheet")
                    break
                local_val = row[en_col].value

                if google_val is None:
                    google_val = ''
                if local_val is None:
                    local_val = ''

                #print(repr(google_val), repr(local_val))

                if str(google_val) != str(local_val):
                    # If ingame col is different, sync local changes to google sheet
                    if en_col == ingame_en_col:
                        label = google_values[i][en_col].label
                        google_worksheet.update_cell(label, local_val)
                    # If the kuoushi col has changed, sync google changes to local sheet
                    elif en_col == original_en_col:
                        cell = local_worksheet.cell(row=i+1, column=en_col+1)
                        cell.value = google_val

    local.workbook.save(local_filename)
"""